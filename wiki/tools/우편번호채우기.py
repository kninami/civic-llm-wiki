#!/usr/bin/env python3
"""주소록에 우편번호를 채워 넣는다. 네이버 검색 결과에서 우편번호를 찾는 방식.

사용법:
    python3 우편번호채우기.py 주소록.xlsx
    python3 우편번호채우기.py 주소록.xlsx --열 "집주소"   # 주소 열을 직접 지정
    python3 우편번호채우기.py 주소록.csv  --개수 10       # 앞 10건만 (검증용)

원본 파일은 건드리지 않고 "<원본이름>_우편번호.xlsx"를 새로 만든다.
자세한 설명과 주의사항은 같은 폴더의 '우편번호-자동검색-파이썬.md' 참조.
"""

import argparse
import csv
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

# Accept 헤더가 없으면 네이버가 403으로 막는다. User-Agent만으로는 부족.
HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"),
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9",
}

# 네이버 주소 카드에만 나타나는 두 가지 형태. 본문·블로그의 "우편번호 12345" 같은
# 예시 문구를 잘못 집지 않도록 느슨한 근접 매칭 대신 구조를 그대로 요구한다.
PATTERNS = [
    re.compile(r"우편번호</span>(\d{5})"),
    re.compile(r"우편번호 · (\d{5}) · 복사"),
]

NOT_FOUND = "#못찾음"
FETCH_FAIL = "#조회실패"


def find_postcode(address, timeout=15):
    """주소 하나의 우편번호를 찾는다. 실패하면 #으로 시작하는 표시를 돌려준다."""
    url = ("https://search.naver.com/search.naver?query="
           + urllib.parse.quote(address))
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as res:
            html = res.read().decode("utf-8", errors="replace")
    except Exception as e:
        print(f"    (조회 실패: {e})", file=sys.stderr)
        return FETCH_FAIL

    for pattern in PATTERNS:
        m = pattern.search(html)
        if m:
            return m.group(1)
    return NOT_FOUND


def load(path):
    """엑셀/CSV를 (헤더, 행목록)으로 읽는다."""
    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        return rows[0], rows[1:]

    import openpyxl  # 엑셀일 때만 필요
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = [["" if c is None else str(c) for c in r]
            for r in ws.iter_rows(values_only=True)]
    return rows[0], rows[1:]


def save(path, header, rows):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


def detect_address_column(header, rows):
    """열 이름에 '주소'가 들어간 열을 찾는다. 없으면 값이 가장 주소처럼 생긴 열."""
    for i, name in enumerate(header):
        if "주소" in str(name):
            return i
    hints = ("시", "군", "구", "로", "길", "동", "읍", "면")
    best, best_score = None, 0
    for i in range(len(header)):
        vals = [r[i] for r in rows[:20] if i < len(r) and r[i]]
        if not vals:
            continue
        score = sum(1 for v in vals if any(h in str(v) for h in hints)) / len(vals)
        if score > best_score:
            best, best_score = i, score
    return best if best_score >= 0.6 else None


def main():
    ap = argparse.ArgumentParser(description="주소록에 우편번호를 채웁니다.")
    ap.add_argument("파일", help="주소록 .xlsx 또는 .csv")
    ap.add_argument("--열", help="주소가 든 열 이름 (생략하면 자동 탐지)")
    ap.add_argument("--개수", type=int, help="앞에서 N건만 처리 (검증용)")
    ap.add_argument("--간격", type=float, default=1.5, help="요청 간 대기 초 (기본 1.5)")
    args = ap.parse_args()

    src = Path(getattr(args, "파일"))
    if not src.exists():
        sys.exit(f"파일이 없습니다: {src}")

    header, rows = load(src)

    col_name = getattr(args, "열")
    if col_name:
        if col_name not in header:
            sys.exit(f"'{col_name}' 열이 없습니다. 있는 열: {', '.join(header)}")
        col = header.index(col_name)
    else:
        col = detect_address_column(header, rows)
        if col is None:
            sys.exit(f"주소 열을 찾지 못했습니다. --열 로 지정하세요. 있는 열: {', '.join(header)}")
        print(f"주소 열: '{header[col]}' (자동 탐지)")

    limit = getattr(args, "개수")
    targets = rows[:limit] if limit else rows
    print(f"대상 {len(targets)}건 · 요청 간격 {getattr(args, '간격')}초 "
          f"· 예상 {len(targets) * getattr(args, '간격') / 60:.1f}분\n")

    cache, results = {}, []
    counts = {"성공": 0, "못찾음": 0, "조회실패": 0}
    for n, row in enumerate(targets, 1):
        addr = str(row[col]).strip() if col < len(row) else ""
        if not addr:
            results.append("")
            continue
        if addr in cache:
            code = cache[addr]
        else:
            code = find_postcode(addr)
            cache[addr] = code
            time.sleep(getattr(args, "간격"))
        results.append(code)

        if code == NOT_FOUND:
            counts["못찾음"] += 1
        elif code == FETCH_FAIL:
            counts["조회실패"] += 1
        else:
            counts["성공"] += 1
        print(f"[{n}/{len(targets)}] {addr[:40]:<40} → {code}")

    out_rows = []
    for i, row in enumerate(rows):
        row = list(row) + [""] * (len(header) - len(row))
        out_rows.append(row + [results[i] if i < len(results) else ""])

    # 검증 모드 결과가 전체 실행 결과를 덮어쓰지 않도록 파일명을 분리한다
    suffix = "_우편번호_샘플" if limit else "_우편번호"
    dst = src.with_name(f"{src.stem}{suffix}.xlsx")
    save(dst, list(header) + ["우편번호"], out_rows)

    print(f"\n{'-' * 46}")
    print(f"성공 {counts['성공']} · 못찾음 {counts['못찾음']} · 조회실패 {counts['조회실패']}")
    print(f"저장: {dst}")
    if counts["못찾음"] or counts["조회실패"]:
        print(f"\n{NOT_FOUND} = 네이버에 우편번호가 안 뜨는 주소 (오래된 지번주소·신축·오타)")
        print(f"{FETCH_FAIL} = 네트워크 오류 또는 네이버 차단. 잠시 뒤 다시 시도")
    print("\n※ 결과를 그대로 믿지 마세요. 최소 10건은 손으로 대조한 뒤 사용하세요.")


if __name__ == "__main__":
    main()
